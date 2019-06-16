from django.shortcuts import render
from snu_moyeo.serializers import MeetingSerializer, SnuUserSerializer, ParticipateSerializer, CommentSerializer, ReportSerializer
from rest_framework import generics
from django.contrib.auth.models import User
from snu_moyeo.models import SnuUser, Meeting, Participate, Comment, Report
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.mail import send_mail
from rest_framework import mixins, filters
from rest_framework.authtoken.models import Token
from django.utils.crypto import get_random_string
from rest_framework import permissions
from django.http import HttpResponse, JsonResponse
from snu_moyeo.permissions import UserOnlyAccess, LeaderOnlyControl, SuperUserOnlyAccess
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
from sdk.api.message import Message
from sdk.exceptions import CoolsmsException
import random
import sys
import django
import urllib.request
from urllib import parse
import json
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Alignment

OPEN = 0
CLOSED = 1
RE_OPEN = 2
RE_CLOSED = 3
BREAK_UP = 4

def send_message(to_number, code):
    # set api key, api secret
    api_key = "NCSTXGIWAWFV3UHU"
    api_secret = "4DO7XGNVKEOGGDDDDXWESGCVNW9MFBGP"

    ## 4 params(to, from, type, text) are mandatory. must be filled
    params = dict()
    params['type'] = 'sms' # Message type ( sms, lms, mms, ata )
    params['to'] =  str(to_number)  # Recipients Number '01000000000, 01000000001'
    params['from'] = '01040079493' # Sender number
    params['text'] = str(code)  # Authentication code

    cool = Message(api_key, api_secret)
    try:
        response = cool.send(params)
        print("Success Count : %s" % response['success_count'])
        print("Error Count : %s" % response['error_count'])
        print("Group ID : %s" % response['group_id'])

        if "error_list" in response:
            print("Error List : %s" % response['error_list'])

    except CoolsmsException as e:
        print("Error Code : %s" % e.code)
        print("Error Message : %s" % e.msg)

class SignUp(mixins.ListModelMixin, mixins.CreateModelMixin, generics.GenericAPIView):
    queryset = SnuUser.objects.all()
    serializer_class = SnuUserSerializer

    # for debugging
    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    def post(self, request, format = None):
        user_serializer = SnuUserSerializer(data = request.data)
        if user_serializer.is_valid():
            user_serializer.save()
            return Response(status.HTTP_201_CREATED)
        else :
            return Response(user_serializer.errors, status = status.HTTP_400_BAD_REQUEST)

class SendEmail(mixins.ListModelMixin, mixins.CreateModelMixin, generics.GenericAPIView):
    queryset = SnuUser.objects.all()
    serializer_class = SnuUserSerializer

    def get(self, request, email):
        user = request.user
        email_token = random.randint(10000000, 99999999)
        send_mail('SnuMoyeo Authenticate', 'Authentication Code : ' + str(email_token),
                'toro.8906@gmail.com', [email], fail_silently = False)

        user_serializer = SnuUserSerializer(user, data = {'mySNU_verification_token':email_token}, partial = True)
        if user_serializer.is_valid():
            user_serializer.save()
            return Response(status = status.HTTP_201_CREATED)
        else :
            return Response(user_serializer.errors, status = status.HTTP_400_BAD_REQUEST)

class EmailAuthenticate (APIView):
    def get(self, request, email_token, email):
        print('Email authenticationg...')
        user = request.user
        if (user.mySNU_verified):
            return Response({'details':'Already verified'}, status = status.HTTP_400_BAD_REQUEST)
        if (user.mySNU_verification_token == email_token):
            user_serializer = SnuUserSerializer(user, data = {'email':email, 'mySNU_verified':True}, partial = True)

        if user_serializer.is_valid():
            user_serializer.save()
            return HttpResponse("Email verification done!", status = 202)
        else :
            return Response(user_serializer.errors, status = status.HTTP_400_BAD_REQUEST)

class SendPhone(mixins.ListModelMixin, mixins.CreateModelMixin, generics.GenericAPIView):
    queryset = SnuUser.objects.all()
    serializer_class = SnuUserSerializer

    def get(self, request, phone_number):
        user = request.user
        phone_token = random.randint(10000000, 99999999)
        send_message(phone_number, phone_token)

        user_serializer = SnuUserSerializer(user, data = {'phone_verification_token':phone_token}, partial = True)
        if user_serializer.is_valid():
            user_serializer.save()
            return Response(status = status.HTTP_201_CREATED)
        else :
            return Response(user_serializer.errors, status = status.HTTP_400_BAD_REQUEST)

class SMSAuthenticate (APIView):
    def get(self, request, phone_token, phone_number) :
        print('Phone authenticationg...')
        user = request.user
        if (user.phone_verified):
            return Response({'details':'Already verified'}, status = status.HTTP_400_BAD_REQUEST)
        if (user.phone_verification_token == phone_token):
            user_serializer = SnuUserSerializer(user, data = {'phone_number':phone_number, 'phone_verified':True}, partial = True)

        if user_serializer.is_valid():
            user_serializer.save()
            return HttpResponse("Phone verification done!", status = 202)
        else :
            return Response(user_serializer.errors, status = status.HTTP_400_BAD_REQUEST)

class LogIn(APIView):
    queryset = SnuUser.objects.all()
    serializer_class = SnuUserSerializer
    permission_classes = ()

    def get(self, request, format = None):
        print('Log in..')
        user = request.user
        if user.mySNU_verified == True and user.phone_verified == True:
            return Response(data = {
                'user_id':user.id,
                'email':user.email,
                'phone_number':user.phone_number,
                'mySNU_verification_token':user.mySNU_verification_token,
                'phone_verification_token':user.phone_verification_token,
                'name':user.name,
                'point':user.point
            }, status = status.HTTP_202_ACCEPTED)
        else :
            return Response(data = {
                'user_id':user.id,
                'email':user.email,
                'phone_number':user.phone_number,
                'mySNU_verified':user.mySNU_verified,
                'phone_verified':user.phone_verified,
                'mySNU_verification_token':user.mySNU_verification_token,
                'phone_verification_token':user.phone_verification_token,
                'name':user.name,
                'point':user.point
            }, status = status.HTTP_403_FORBIDDEN)

class SnuUserList(generics.ListAPIView):
    queryset = SnuUser.objects.all()
    serializer_class = SnuUserSerializer

class SnuUserDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = SnuUser.objects.all()
    serializer_class = SnuUserSerializer
    permission_classes = (UserOnlyAccess,)

def changeState():
    current = django.utils.timezone.now()
    print(current)

    # Need optimization !! (to Jongmin)
    for meeting in Meeting.objects.all() :
        if meeting.due < current and meeting.state == OPEN :
            if meeting.min_people > meeting.members.all().count() :
                meeting.state = BREAK_UP
            else :
                meeting.state = CLOSED
            meeting.save()
    '''
    Meeting.objects.filter(Q(state = OPEN) & Q(due__lt = current)).update(state = CLOSED)
    for meeting_object in Meeting.objects.filter(Q(state = CLOSED) & Q(due__lt = current)):
        meeting_object.save()

    Meeting.objects.filter(Q(state = OPEN) & Q(due__lt = current)).update(state = CLOSED)
    '''


class MeetingList(generics.ListCreateAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    permission_classes = (permissions.IsAuthenticatedOrReadOnly,)
    def perform_create(self, serializer):
        serializer.save(leader = self.request.user)

    def get(self, request, *args, **kwargs):
        changeState()
        return self.list(request, *args, **kwargs)

class MeetingDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, LeaderOnlyControl,)

class ParticipateList(generics.ListCreateAPIView):
    queryset = Participate.objects.all()
    serializer_class = ParticipateSerializer


class ParticipateDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Participate.objects.all()
    serializer_class = ParticipateSerializer

    def delete(self, request, *args, **kwargs) :
        outer_id = self.get_object().user_id.id
        print((outer_id))
        outer = SnuUser.objects.get(pk = outer_id)
        target_meetingid = self.get_object().meeting_id.id
        target_meeting = Meeting.objects.get(pk = target_meetingid)
        # meeting_min = target_meeting.min_people
        meeting_state = target_meeting.state
        
        if meeting_state == CLOSED :
            outer.point += 3
            outer.save()

        return self.destroy(request, *args, **kwargs)


def get_participate(request, in_userid, in_meetingid):
    participate_obj1 = Participate.objects.filter(Q(user_id_id = in_userid))
    participate_list1 = participate_obj1.values()
    participate_obj2 = participate_obj1.filter(Q(meeting_id_id = in_meetingid))
    participate_list2 = participate_obj2.values()

    if len(participate_list2) == 0 :
        return HttpResponse({}, status = status.HTTP_404_NOT_FOUND)
    participate_id = participate_list2[0]['id']
    return HttpResponse(participate_id, status = status.HTTP_200_OK)

class RecentList(generics.ListAPIView):
    queryset = Meeting.objects.all().filter(Q(state = OPEN) | Q(state = RE_OPEN))[:5]
    serializer_class = MeetingSerializer

    def get(self, request, *args, **kwargs):
        changeState()
        return self.list(request, *args, **kwargs)

class ImpendingList(generics.ListAPIView):
    queryset = Meeting.objects.all().filter(Q(state = OPEN) | Q(state = RE_OPEN)).order_by('due')[:5]
    serializer_class = MeetingSerializer

    def get(self, request, *args, **kwargs):
        changeState()
        return self.list(request, *args, **kwargs)

class LeadList(generics.ListAPIView):
    serializer_class = MeetingSerializer

    def get_queryset(self):
        user = self.request.user
        if (not user.is_anonymous):
            lead_user = SnuUser.objects.get(id = user.id)
            return lead_user.lead_meeting.all()
        return Meeting.objects.none()

    def get(self, request, *args, **kwargs):
        changeState()
        return self.list(request, *args, **kwargs)

class JoinList (generics.ListAPIView):
    serializer_class = MeetingSerializer

    def get_queryset(self):
        user = self.request.user
        if (not user.is_anonymous):
            user_id = user.id
            join_user = SnuUser.objects.get(id = user_id)
            return join_user.meetings.all().filter(~Q(state = BREAK_UP))
        return Meeting.objects.none()

    def get(self, request, *args, **kwargs):
        changeState()
        return self.list(request, *args, **kwargs)

class HistoryList (generics.ListAPIView):
    serializer_class = MeetingSerializer

    def get_queryset(self):
        user = self.request.user
        if (not user.is_anonymous):
            user_id = user.id
            history_user = SnuUser.objects.get(id = user_id)
            print(history_user)
            print(history_user.meetings.all())
            return history_user.meetings.all().filter(Q(state = BREAK_UP))
        return Meeting.objects.none()

    def get(self, request, *args, **kwargs):
        changeState()
        return self.list(request, *args, **kwargs)

class CustomPagination(PageNumberPagination):
    page_size = 3
    page_size_query_param = 'page_size'
    max_page_size = 3

    def get_paginated_response(self, data):
        return Response({
            'links': {
                'next': self.get_next_link(),
                'previous': self.get_previous_link()
                },
            'count': self.page.paginator.count,
            'page_size': self.page_size,
            'results': data
            })

'''
simple version pagination (not used now)
class TempList(generics.ListAPIView):
    serializer_class = MeetingSerializer
    pagination_class = CustomPagination

    def get_queryset(self):
        temp_meeting = Meeting.objects.filter(~Q(state=4))
        return temp_meeting
'''

class ListView(APIView):
    def get(self, request, in_kind, format = None):
        changeState()
        kind_meeting = Meeting.objects.filter(Q(kind = in_kind) & (Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(kind_meeting, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)

class CommentList(generics.ListCreateAPIView):
    queryset = Comment.objects.all()
    serializer_class = CommentSerializer

    def perform_create(self, serializer):
        serializer.save(writer = self.request.user)

class CommentDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Comment.objects.all()
    serializer_class = CommentSerializer
    permission_classes = () # (UserOnlyAccess,) : 에러 발생

class CommentOnMeeting(APIView):
    def get(self, request, in_meetingid):
        comments = Comment.objects.filter(Q(meetingid_id = in_meetingid))
        serializer = CommentSerializer(comments, many = True)
        return Response(serializer.data)


class MeetingSearchAllView(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter((Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)



class MeetingSearch0View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=0)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)

class MeetingSearch1View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=1)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)


class MeetingSearch2View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=2)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)


class MeetingSearch3View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=3)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)


class MeetingSearch4View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        print('hi')
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=4)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)


class MeetingSearch5View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=5)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)


class MeetingSearch6View(generics.ListAPIView):
    queryset = Meeting.objects.all()
    serializer_class = MeetingSerializer
    filter_backends = (filters.SearchFilter,)
    search_fields = ('title', 'description')
    def get(self, request):
        changeState()
        self.filter_backends = (filters.SearchFilter,)
        self.search_fields = ('title', 'description',)
        queryset = self.filter_queryset(self.get_queryset()).filter(Q(kind=6)&(Q(state = OPEN) | Q(state = RE_OPEN)))
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(queryset, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)





class ReportList(generics.ListCreateAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer

    def perform_create(self, serializer):
        serializer.save(reporter = self.request.user)

    def get_queryset(self):
        user = self.request.user
        if (user.is_superuser):
            return Report.objects.all()
        return Meeting.objects.none()


    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)


class ReportDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    def get_queryset(self):
        user = self.request.user
        if (user.is_superuser):
            return Report.objects.all()
        return Meeting.objects.none()
    '''
    def delete(self, request, *args, **kwargs) :
        if self.get_object().isHandled == True :
            target = SnuUser.objects.get(pk = self.get_object().reporteeid)
            target_point = self.get_object().point
            target.point -= target_point
            target.save()
        return self.destroy(request, *args, **kwargs)
    '''

def InfoExcel(request, meeting_id) :
    write_wb = Workbook()
    write_ws = write_wb.active


    box = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )




    write_ws['A1'] = '이름'
    write_ws['A1'].border = box
    write_ws['A1'].font = Font(size=11, bold=True)
    write_ws.column_dimensions['A'].width = 15
    
    write_ws['B1'] = '닉네임'
    write_ws['B1'].border = box
    write_ws['B1'].font = Font(size=11, bold=True)
    write_ws.column_dimensions['B'].width = 15


    write_ws['C1'] = '전화번호'
    write_ws['C1'].border = box
    write_ws['C1'].font = Font(size=11, bold=True)
    write_ws.column_dimensions['C'].width = 20

    write_ws['D1'] = '이메일'
    write_ws['D1'].border = box
    write_ws['D1'].font = Font(size=11, bold=True)
    write_ws.column_dimensions['D'].width = 30

    write_ws['E1'] = '벌점'
    write_ws['E1'].border = box
    write_ws['E1'].font = Font(size=11, bold=True)

    meeting_info = Meeting.objects.get(pk = meeting_id)
    participants = meeting_info.members.all()
    
    ind = 2
    for user in participants :
        write_ws.cell(row = ind, column = 1).value = user.name
        write_ws.cell(row = ind, column = 1).border = box
        write_ws.cell(row = ind, column = 2).value = user.username
        write_ws.cell(row = ind, column = 2).border = box
        write_ws.cell(row = ind, column = 3).value = user.phone_number
        write_ws.cell(row = ind, column = 3).border = box
        write_ws.cell(row = ind, column = 4).value = user.email
        write_ws.cell(row = ind, column = 4).border = box
        write_ws.cell(row = ind, column = 5).value = user.point
        write_ws.cell(row = ind, column = 5).border = box
        ind += 1

    title = str(meeting_info.title).replace(' ','_')
    excel_filename = 'media/' + str(title)  + '_모임참여자 정보.xlsx'
    write_wb.save(excel_filename)

    file_location = excel_filename
    try:
        with open(file_location, 'rb') as f:
           file_data = f.read()

        # sending response
        response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="member_info.xlsx"'

    except IOError:
        # handle file not exist case here
        response = HttpResponseNotFound('<h1>File not exist</h1>')

    return response
    '''
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + excel_filename
    return response
    '''

def searchShop(request, search_word) :
    client_id = "N6c7MAUvz7uiaMUNt1Ww" # 애플리케이션 등록시 발급 받은 값 입력
    client_secret = "1hdjaYpmI6" # 애플리케이션 등록시 발급 받은 값 입력
    encText = urllib.parse.quote(search_word)
    print(encText)
    url = "https://openapi.naver.com/v1/search/shop?query=" + encText +"&display=3&sort=sim"
    request = urllib.request.Request(url)
    print(request)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if(rescode==200):
        response_body = response.read()
        #print(response_body.decode('utf-8'))
        json_dict_def = json.loads(response_body.decode('utf-8'))
        json_dict = json_dict_def
        items = json_dict['items']
        ind = 0
        for item in items :
            replace_tag = item['title'].replace('<b>','').replace('</b>','')
            item['title'] = replace_tag
            json_dict['items'][ind] = item
            ind +=1
        #print(type(json_dict))
        #print(json_dict)
        #print(json_dict)
        return JsonResponse(json_dict)
    else:
        print("Error Code:" + rescode)
        return HttpResponse({}, status = status.HTTP_404_NOT_FOUND)


def captcha(request):
    client_id = "N6c7MAUvz7uiaMUNt1Ww" # 애플리케이션 등록시 발급 받은 값 입력
    client_secret = "1hdjaYpmI6" # 애플리케이션 등록시 발급 받은 값 입력

    url = "https://openapi.naver.com/v1/captcha/nkey?code=0"
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)

    rescode = response.getcode()
    capch_key = ''
    if(rescode==200):
        response_body = response.read()
        capch_key = response_body.decode('utf-8')
        print((capch_key))
    else:
        print("Error Code:" + rescode)


    url = "https://openapi.naver.com/v1/captcha/ncaptcha.bin?key=" + capch_key.split('\"')[3]
    print(url)
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)

    rescode = response.getcode()
    if(rescode==200):
        print("캡차 이미지 저장")
        response_body = response.read()
        filename = 'media/captcha'
        filename = filename + str(random.randint(1, 100))+'.jpg'
        with open(filename, 'wb') as f:
            f.write(response_body)
        #return HttpResponse(response_body)
    else:
        print('wrong')

    json_dict = {'image': filename,'key': capch_key.split('\"')[3]}
    return JsonResponse(json_dict)

def captcha_verify(request, in_key, in_value) :
    client_id = "N6c7MAUvz7uiaMUNt1Ww" # 애플리케이션 등록시 발급 받은 값 입력
    client_secret = "1hdjaYpmI6" # 애플리케이션 등록시 발급 받은 값 입력

    code = "1"
    key = in_key
    value = in_value
    url = "https://openapi.naver.com/v1/captcha/nkey?code=" + code + "&key=" + key + "&value=" + value
    print(url)
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if(rescode==200):
        response_body = response.read()
        return HttpResponse(response_body)
        #print(response_body.decode('utf-8'))
    else:
        print("Error Code:" + rescode)


'''
class ListView(APIView):
    def get(self, request, search):
        changeState()
        meeting = Meeting.objects.filter(
        paginator = CustomPagination()
        result_page = paginator.paginate_queryset(meeting, request)
        serializer = MeetingSerializer(result_page, many = True)
        return paginator.get_paginated_response(serializer.data)
'''
'''
def get_comments_on_meeting(request, in_meetingid):
    temp = Meeting.objects.all()
    return HttpResponse(temp.values(),status=200)
    comments = Comment.objects.filter(Q(meeting_id_id = in_meetingid))
    return HttpResponse(comments.values(),status = 200)
    serializer = CommentSerializer(data = comments, many = True)
    if serializer.is_valid() :
        return HttpResponse(serializer.data, status = status.HTTP_200_OK)
    else :
        return HttpResponse(serializer.errors, status = 400)

def get_participate(request, in_userid, in_meetingid):
    participate_obj1 = Participate.objects.filter(Q(user_id_id = in_userid))
    participate_list1 = participate_obj1.values()
    participate_obj2 = participate_obj1.filter(Q(meeting_id_id = in_meetingid))
    participate_list2 = participate_obj2.values()
    # print(participate_obj2.values())

    if len(participate_list2) == 0 :
        return HttpResponse({}, status = status.HTTP_404_NOT_FOUND)
    participate_id = participate_list2[0]['id']
    return HttpResponse(participate_id, status = status.HTTP_200_OK)
'''
